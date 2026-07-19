"""Chayabithi Cafe – Cashier SMS → Excel updater (Gradio app).

Paste the cashier's daily Bengali sales SMS, upload the monthly workbook, and
download the workbook with that day's column filled in. After download the
generated file is wiped from the server.

Parsing is done offline by default (fast, free, reliable). An optional Groq LLM
engine is available via the UI toggle / environment variables.
"""

import os
import uuid
import shutil
import datetime
from pathlib import Path

import gradio as gr
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator

import parser_core as pc

# ZeroGPU (free-tier Spaces): at least one @spaces.GPU is required at startup.
# Effect-free on CPU / local runs when the package is missing.
try:
    import spaces
except ImportError:  # local / non-ZeroGPU
    class spaces:  # type: ignore
        @staticmethod
        def GPU(fn=None, **_kwargs):
            if fn is None:
                return lambda f: f
            return fn

APP_DIR = Path(__file__).parent.resolve()
TEMP_DIR = APP_DIR / "temp"
TEMP_DIR.mkdir(exist_ok=True)

GROQ_API_KEY = os.getenv("GROQ_API_KEY", "").strip()
# llama-3.1-70b-versatile is DECOMMISSIONED on Groq (and llama-3.3-70b is being
# retired 2026-08-16). Default to a current, supported model.
GROQ_MODEL = os.getenv("GROQ_MODEL", "openai/gpt-oss-120b").strip()


# ---------------------------------------------------------------------------
# Workbook writing
# ---------------------------------------------------------------------------
def _find_header_row(ws):
    for r in range(1, min(ws.max_row, 30) + 1):
        if str(ws.cell(r, 1).value).strip() == "SL No":
            return r
    return 9


def _find_date_column(ws, header_row, date):
    """date = (day, month, year|None). Match against datetime header cells."""
    day, month, year = date
    fallback = None
    for c in range(3, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, datetime.datetime):
            if v.day == day and v.month == month:
                if year and v.year != year:
                    continue
                return c
        elif v is not None:
            sv = str(v).strip()
            if sv and (f"{day}" in sv and f"{month}" in sv):
                fallback = fallback or c
    return fallback


def _build_sl_rows(ws):
    """Map SL number -> (name_row, qty_row, rate_row, total_row)."""
    mp = {}
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 2).value).strip() == "Item name":
            sl = ws.cell(r, 1).value
            try:
                sl = int(sl)
            except (TypeError, ValueError):
                continue
            mp[sl] = (r, r + 1, r + 2, r + 3)
    return mp


def _find_jumping_rows(ws):
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 1).value).strip() == "Jumping":
            return {"taka": r, "person": r + 1, "total": r + 2}
    return None


def _num(v):
    """Return int when the value is whole, else float."""
    f = float(v)
    return int(f) if f == int(f) else f


def _pick_ref_col(ws, header_row):
    """A date column that already has its formulas filled, to copy from."""
    for c in range(3, ws.max_column + 1):
        if c == ws.max_column:  # skip the "Total" summary column
            continue
        for r in range(header_row, ws.max_row + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.startswith("="):
                return c
    return None


def _fill_formulas(ws, col, header_row):
    """Copy every formula (item totals, day Total, Grand Total, Per Customer…)
    from a reference date column into ``col``, translating cell references.

    This mirrors exactly how the other days are computed, so the new day's
    daily total and grand total appear automatically."""
    ref_col = _pick_ref_col(ws, header_row)
    if ref_col is None:
        return
    ref_letter = get_column_letter(ref_col)
    col_letter = get_column_letter(col)
    for r in range(1, ws.max_row + 1):
        ref = ws.cell(r, ref_col).value
        if not (isinstance(ref, str) and ref.startswith("=")):
            continue
        if ws.cell(r, col).value not in (None, ""):
            continue
        label = str(ws.cell(r, 2).value or "").strip().lower()
        # "Per Customer" = Total / Customer-Num. Skip it when the customer
        # count for this day is missing, to avoid a #DIV/0! error.
        if "per customer" in label and ws.cell(r - 1, col).value in (None, ""):
            continue
        ws.cell(r, col).value = Translator(
            ref, origin=f"{ref_letter}{r}"
        ).translate_formula(f"{col_letter}{r}")


def _copy_label(ws, row, col):
    """Copy a repeated text label (e.g. the item name) from the first filled
    date column into ``col`` when it is empty."""
    if ws.cell(row, col).value not in (None, ""):
        return
    for cc in range(3, ws.max_column):  # skip the trailing "Total" column
        if cc == col:
            continue
        v = ws.cell(row, cc).value
        if v not in (None, ""):
            ws.cell(row, col).value = v
            return


def update_workbook(xlsx_path: str, records, jumping, date, customer_num=None):
    wb = load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]

    header_row = _find_header_row(ws)
    col = _find_date_column(ws, header_row, date)
    if col is None:
        raise gr.Error(
            f"তারিখ {date[0]:02d}/{date[1]:02d} শীটের হেডারে পাওয়া যায়নি। "
            "ফাইলটি সঠিক মাসের কিনা যাচাই করুন।"
        )
    col_letter = get_column_letter(col)

    sl_rows = _build_sl_rows(ws)
    written, skipped = [], []

    for rec in records:
        rows = sl_rows.get(rec["sl"])
        if not rows:
            skipped.append(rec["sl"])
            continue
        name_row, qty_row, rate_row, total_row = rows
        _copy_label(ws, name_row, col)  # ensure the item name shows in this column
        ws.cell(qty_row, col).value = _num(rec["qty"])
        if ws.cell(rate_row, col).value in (None, "") and rec.get("rate"):
            ws.cell(rate_row, col).value = _num(rec["rate"])
        written.append(rec["sl"])

    if jumping:
        jr = _find_jumping_rows(ws)
        if jr:
            ws.cell(jr["person"], col).value = _num(jumping["person"])
            if ws.cell(jr["taka"], col).value in (None, "") and jumping.get("taka"):
                ws.cell(jr["taka"], col).value = _num(jumping["taka"])

    # Optional customer count -> lets "Per Customer" compute below.
    if customer_num not in (None, ""):
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 2).value).strip() == "Customer Num":
                ws.cell(r, col).value = _num(customer_num)
                break

    # Fill item-total, daily-Total, Grand-Total (and Per Customer) formulas by
    # translating them from an already-filled day column.
    _fill_formulas(ws, col, header_row)

    # The temp dir can be wiped by a page refresh (unload) or the clear button,
    # so make sure it exists right before saving.
    TEMP_DIR.mkdir(parents=True, exist_ok=True)
    out_path = TEMP_DIR / f"updated_{uuid.uuid4().hex}.xlsx"
    wb.save(out_path)
    return str(out_path), col_letter, written, skipped


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def _get_groq_client():
    if not GROQ_API_KEY:
        return None
    try:
        from groq import Groq
        return Groq(api_key=GROQ_API_KEY)
    except Exception:
        return None


@spaces.GPU(duration=60)
def process(sms_text, xlsx_file, engine, customer_num=None):
    if xlsx_file is None:
        raise gr.Error("অনুগ্রহ করে মাসিক XLSX ফাইলটি আপলোড করুন।")
    if not (sms_text or "").strip():
        raise gr.Error("অনুগ্রহ করে ক্যাশিয়ারের এসএমএস টেক্সট পেস্ট করুন।")

    used_engine = "offline"
    warnings = []
    if engine == "AI (Groq LLM)":
        client = _get_groq_client()
        if client is None:
            warnings.append("GROQ_API_KEY নেই/ভুল — অফলাইন পার্সারে ফিরে গেছি।")
            records, jumping, date, w = pc.parse_sms(sms_text)
        else:
            try:
                records, jumping, date, w = pc.llm_parse_sms(sms_text, client, GROQ_MODEL)
                used_engine = f"Groq · {GROQ_MODEL}"
            except Exception as e:
                warnings.append(f"LLM ব্যর্থ ({type(e).__name__}) — অফলাইন পার্সার ব্যবহার করা হলো।")
                records, jumping, date, w = pc.parse_sms(sms_text)
        warnings += w
    else:
        records, jumping, date, warnings = pc.parse_sms(sms_text)

    if date is None:
        raise gr.Error("মেসেজে তারিখ পাওয়া যায়নি (যেমন: তারিখ: ১৫/০৭/২৬)।")
    if not records:
        raise gr.Error("কোনো আইটেম শনাক্ত করা যায়নি। মেসেজ ফরম্যাট যাচাই করুন।")

    out_path, col_letter, written, skipped = update_workbook(
        xlsx_file.name, records, jumping, date, customer_num
    )

    rows = []
    for rec in sorted(records, key=lambda x: x["sl"]):
        rows.append([
            rec["sl"], pc.SL_TO_EN.get(rec["sl"], "?"),
            _num(rec["qty"]), _num(rec["rate"]),
        ])
    if jumping:
        rows.append(["—", "Jumping (person)", _num(jumping["person"]), _num(jumping["taka"])])

    summary = (
        f"ইঞ্জিন: {used_engine}\n"
        f"তারিখ: {date[0]:02d}/{date[1]:02d}/{date[2] or ''}  →  কলাম {col_letter}\n"
        f"লেখা হয়েছে: {len(written)} আইটেম" + (f" · বাদ: {skipped}" if skipped else "")
    )
    if warnings:
        summary += "\n\n⚠️ " + "\n⚠️ ".join(warnings)

    return summary, rows, out_path


def clear_backend():
    """Wipe all generated files from the server."""
    if TEMP_DIR.exists():
        shutil.rmtree(TEMP_DIR, ignore_errors=True)
    TEMP_DIR.mkdir(exist_ok=True)
    return "✅ ব্যাকএন্ড পরিষ্কার করা হয়েছে।", None, None


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------
SAMPLE = "তারিখ: ১৫/০৭/২৬\n(এখানে ক্যাশিয়ারের মেসেজ পেস্ট করুন)"

with gr.Blocks(title="Chayabithi Cafe — SMS to Excel", analytics_enabled=False) as demo:
    gr.Markdown(
        "# ☕ ছায়াবীথি ক্যাফে — ক্যাশিয়ার SMS → Excel\n"
        "ক্যাশিয়ারের দৈনিক মেসেজ পেস্ট করুন, মাসিক XLSX আপলোড করুন, "
        "আর সেই দিনের কলাম পূরণ করা আপডেটেড ফাইল ডাউনলোড করুন।"
    )
    with gr.Row():
        with gr.Column(scale=1):
            sms = gr.Textbox(lines=20, label="ক্যাশিয়ারের SMS", placeholder=SAMPLE)
            xlsx = gr.File(label="মাসিক XLSX আপলোড", file_types=[".xlsx"])
            customer_num = gr.Number(
                label="গ্রাহক সংখ্যা (Customer number) — ঐচ্ছিক",
                precision=0,
                value=None,
            )
            engine = gr.Radio(
                ["Offline (দ্রুত, বিনামূল্যে)", "AI (Groq LLM)"],
                value="Offline (দ্রুত, বিনামূল্যে)",
                label="পার্সিং ইঞ্জিন",
            )
            with gr.Row():
                btn = gr.Button("প্রসেস করুন", variant="primary")
                clear_btn = gr.Button("ব্যাকএন্ড ক্লিয়ার")
        with gr.Column(scale=1):
            summary = gr.Textbox(label="সারসংক্ষেপ", lines=6)
            table = gr.Dataframe(
                headers=["SL", "Item", "Qty", "Rate"],
                label="শনাক্তকৃত বিক্রয়",
                wrap=True,
            )
            out_xlsx = gr.File(label="আপডেটেড XLSX ডাউনলোড")

    btn.click(process, inputs=[sms, xlsx, engine, customer_num], outputs=[summary, table, out_xlsx])
    clear_btn.click(clear_backend, outputs=[summary, table, out_xlsx])
    demo.unload(lambda: shutil.rmtree(TEMP_DIR, ignore_errors=True))

if __name__ == "__main__":
    # SSR can kill the Space process after bind; keep classic Gradio serve.
    demo.launch(ssr_mode=False)
